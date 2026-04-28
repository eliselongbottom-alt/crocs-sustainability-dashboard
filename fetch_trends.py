#!/usr/bin/env python3
"""
fetch_trends.py — Crocs Google Trends Fetcher
Pulls Google Trends data for "Crocs" across 8 markets and writes:
  1. crocs_trends_YYYY-MM-DD.xlsx  (boss's Excel format)
  2. google-trends-data.js         (updates the dashboard automatically)

HOW TO GET AROUND GOOGLE'S BLOCKING:
  Option A (recommended — free 100 searches/month):
    Sign up at serpapi.com, get your API key, then run:
      python3 fetch_trends.py --api serpapi --key YOUR_API_KEY

  Option B (pytrends — free but may get 429 blocked):
    python3 fetch_trends.py --api pytrends
    If blocked: add --cookies (copies cookies from your Chrome session)
    If still blocked: try a VPN or wait a few hours

  Option C (export from Google Trends manually):
    1. Go to trends.google.com, search "Crocs", change geo & timeframe
    2. Download CSV, rename to match a market, repeat for each
    3. Run: python3 fetch_trends.py --from-csv ./csv_exports/

Usage:
  pip install pytrends openpyxl requests
  python3 fetch_trends.py --api pytrends
"""

import sys
import time
import json
import math
import random
import argparse
import datetime
import os

# ── Config ────────────────────────────────────────────────────────────────────
KEYWORD = "Crocs"
MARKETS = [
    ("United States", "US"),
    ("United Kingdom", "GB"),
    ("Germany",        "DE"),
    ("France",         "FR"),
    ("South Korea",    "KR"),
    ("Japan",          "JP"),
    ("India",          "IN"),
    ("China",          "CN"),
]
TIMEFRAMES = [("90d", "today 3-m"), ("30d", "today 1-m")]
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Helpers ───────────────────────────────────────────────────────────────────
def log(msg): print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {msg}")


def sleep_politely(min_s=4, max_s=9):
    """Random delay to avoid triggering Google's rate limiter."""
    delay = random.uniform(min_s, max_s)
    log(f"  waiting {delay:.1f}s …")
    time.sleep(delay)


# ── pytrends fetcher ──────────────────────────────────────────────────────────
def fetch_with_pytrends(market_name, geo, tf_label, tf_google, use_cookies=False):
    try:
        from pytrends.request import TrendReq
    except ImportError:
        print("ERROR: pytrends not installed. Run: pip install pytrends")
        sys.exit(1)

    kwargs = {
        "hl": "en-US",
        "tz": 0,
        "timeout": (10, 25),
        "retries": 3,
        "backoff_factor": 2,
    }

    if use_cookies:
        import http.cookiejar, browser_cookie3
        try:
            cookies = browser_cookie3.chrome(domain_name=".google.com")
            kwargs["requests_args"] = {"cookies": cookies}
            log(f"  Using Chrome cookies for {market_name}")
        except Exception as e:
            log(f"  WARNING: Could not load Chrome cookies ({e}). Proceeding without.")

    pt = TrendReq(**kwargs)

    log(f"  Fetching {market_name} ({geo}) / {tf_label} …")

    try:
        pt.build_payload([KEYWORD], timeframe=tf_google, geo=geo)
        sleep_politely()

        iot = pt.interest_over_time()
        related = pt.related_queries()

        iot_series = []
        if iot is not None and not iot.empty and KEYWORD in iot.columns:
            for idx, row in iot.iterrows():
                iot_series.append({
                    "date": str(idx.date()),
                    "value": int(row[KEYWORD]),
                })

        top_queries = []
        rising_queries = []
        rq = related.get(KEYWORD, {})
        if rq:
            top_df = rq.get("top")
            if top_df is not None and not top_df.empty:
                for _, row in top_df.iterrows():
                    top_queries.append({
                        "term": str(row["query"]),
                        "score": int(row["value"]),
                    })

            rising_df = rq.get("rising")
            if rising_df is not None and not rising_df.empty:
                for _, row in rising_df.iterrows():
                    val = row["value"]
                    is_breakout = (val >= 5000)
                    growth = "Breakout" if is_breakout else f"+{int(val)}%"
                    rising_queries.append({
                        "term": str(row["query"]),
                        "growth": growth,
                        "type": "breakout" if is_breakout else "rising",
                    })

        vals = [p["value"] for p in iot_series]
        return {
            "avgInterest": round(sum(vals) / len(vals)) if vals else None,
            "peakInterest": max(vals) if vals else None,
            "interestOverTime": iot_series,
            "topQueries": top_queries,
            "risingQueries": rising_queries,
        }

    except Exception as e:
        log(f"  ERROR for {market_name}/{tf_label}: {e}")
        return {
            "avgInterest": None,
            "peakInterest": None,
            "interestOverTime": [],
            "topQueries": [],
            "risingQueries": [],
            "error": str(e),
        }


# ── SerpAPI fetcher ───────────────────────────────────────────────────────────
def fetch_with_serpapi(market_name, geo, tf_label, tf_google, api_key):
    import requests

    log(f"  Fetching {market_name} ({geo}) / {tf_label} via SerpAPI …")

    base = "https://serpapi.com/search"
    params = {
        "engine": "google_trends",
        "q": KEYWORD,
        "geo": geo,
        "date": tf_google,
        "api_key": api_key,
    }

    try:
        r = requests.get(base, params=params, timeout=30)
        r.raise_for_status()
        data = r.json()
        sleep_politely(1, 3)

        iot_series = []
        for point in data.get("interest_over_time", {}).get("timeline_data", []):
            iot_series.append({
                "date": point.get("date", "")[:10],
                "value": int(point.get("values", [{}])[0].get("extracted_value", 0)),
            })

        top_queries, rising_queries = [], []
        rq = data.get("related_queries", {})
        for item in rq.get("top", []):
            top_queries.append({"term": item.get("query", ""), "score": int(item.get("value", 0))})
        for item in rq.get("rising", []):
            val_str = item.get("value", "")
            is_breakout = "Breakout" in str(val_str)
            growth = "Breakout" if is_breakout else str(val_str)
            rising_queries.append({
                "term": item.get("query", ""),
                "growth": growth,
                "type": "breakout" if is_breakout else "rising",
            })

        vals = [p["value"] for p in iot_series]
        return {
            "avgInterest": round(sum(vals) / len(vals)) if vals else None,
            "peakInterest": max(vals) if vals else None,
            "interestOverTime": iot_series,
            "topQueries": top_queries,
            "risingQueries": rising_queries,
        }

    except Exception as e:
        log(f"  ERROR for {market_name}/{tf_label}: {e}")
        return {
            "avgInterest": None, "peakInterest": None,
            "interestOverTime": [], "topQueries": [], "risingQueries": [],
            "error": str(e),
        }


# ── Excel writer ──────────────────────────────────────────────────────────────
def write_excel(all_data, pulled_at, output_path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log("WARNING: openpyxl not installed — skipping Excel output. Run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()

    # ── Summary sheet ──
    ws_sum = wb.active
    ws_sum.title = "Summary"
    pulled_str = pulled_at.strftime("%Y-%m-%d %H:%M")
    ws_sum["A1"] = f"Crocs — Google Trends snapshot"
    ws_sum["A2"] = f"Pulled: {pulled_str} | Keyword: {KEYWORD}"
    ws_sum["A3"] = "Note: Trends interest is normalized 0–100 within each market+timeframe — NOT comparable across markets in absolute terms."
    ws_sum.append([])

    headers = ["Market", "Geo", "Timeframe", "Avg interest", "Peak interest", "# Rising terms", "# Breakout terms"]
    ws_sum.append(headers)
    for col_num, h in enumerate(headers, 1):
        cell = ws_sum.cell(row=5, column=col_num)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1D1D1B")
        cell.font = Font(bold=True, color="FFFFFF")

    row = 6
    for market_name, geo in MARKETS:
        for tf_label, _ in TIMEFRAMES:
            d = all_data.get(market_name, {}).get(tf_label)
            if d and not d.get("blocked"):
                ws_sum.cell(row=row, column=1, value=market_name)
                ws_sum.cell(row=row, column=2, value=geo)
                ws_sum.cell(row=row, column=3, value=tf_label)
                ws_sum.cell(row=row, column=4, value=d.get("avgInterest") or "—")
                ws_sum.cell(row=row, column=5, value=d.get("peakInterest") or "—")
                rising = [q for q in d.get("risingQueries", []) if q.get("type") == "rising"]
                breakout = [q for q in d.get("risingQueries", []) if q.get("type") == "breakout"]
                ws_sum.cell(row=row, column=6, value=len(rising))
                ws_sum.cell(row=row, column=7, value=len(breakout))
            else:
                ws_sum.cell(row=row, column=1, value=market_name)
                ws_sum.cell(row=row, column=2, value=geo)
                ws_sum.cell(row=row, column=3, value=tf_label)
                ws_sum.cell(row=row, column=4, value="—")
                ws_sum.cell(row=row, column=5, value="—")
                ws_sum.cell(row=row, column=6, value=0)
                ws_sum.cell(row=row, column=7, value=0)
            row += 1

    ws_sum.append([])
    ws_sum.append([])
    ws_sum.append(["Breakout & top rising terms by market"])
    ws_sum.append(['"Breakout" = >5000% growth, flagged by Google. Otherwise, % is rise vs prior period.'])
    ws_sum.append([])
    ws_sum.append(["Market", "Timeframe", "Term", "Type", "Growth"])
    for market_name, geo in MARKETS:
        for tf_label, _ in TIMEFRAMES:
            d = all_data.get(market_name, {}).get(tf_label) or {}
            rq = d.get("risingQueries", [])
            if not rq:
                ws_sum.append([market_name, tf_label, "(no rising terms returned)", None, None])
            else:
                for q in rq:
                    ws_sum.append([market_name, tf_label, q.get("term"), q.get("type"), q.get("growth")])

    for col in ws_sum.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws_sum.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    # ── Per-market sheets ──
    for market_name, geo in MARKETS:
        ws = wb.create_sheet(title=market_name)
        ws["A1"] = f"{market_name} — {KEYWORD}"
        ws["A2"] = f"Geo: {geo}"
        ws.append([])

        d90 = all_data.get(market_name, {}).get("90d") or {}
        d30 = all_data.get(market_name, {}).get("30d") or {}

        ws.append(["90D — Interest over time", None, None, None, "30D — Interest over time"])
        ws.append(["Date", "Interest", None, None, "Date", "Interest"])
        iot90 = d90.get("interestOverTime", [])
        iot30 = d30.get("interestOverTime", [])
        max_rows = max(len(iot90), len(iot30), 1)
        for i in range(max_rows):
            r90 = iot90[i] if i < len(iot90) else {"date": "(no data)", "value": None}
            r30 = iot30[i] if i < len(iot30) else {"date": None, "value": None}
            ws.append([r90["date"], r90.get("value"), None, None, r30.get("date") or "", r30.get("value")])

        ws.append([])
        ws.append([])
        ws.append(["90D — Top related queries", None, None, None, "30D — Top related queries"])
        ws.append(["Term", "Score", None, None, "Term", "Score"])
        tq90 = d90.get("topQueries", [{"term": "(none)", "score": None}])
        tq30 = d30.get("topQueries", [{"term": "(none)", "score": None}])
        for i in range(max(len(tq90), len(tq30), 1)):
            r90 = tq90[i] if i < len(tq90) else {}
            r30 = tq30[i] if i < len(tq30) else {}
            ws.append([r90.get("term", ""), r90.get("score"), None, None, r30.get("term", ""), r30.get("score")])

        ws.append([])
        ws.append([])
        ws.append(["90D — Rising / Breakout queries", None, None, None, "30D — Rising / Breakout queries"])
        ws.append(["Term", "Growth", None, None, "Term", "Growth"])
        rq90 = d90.get("risingQueries", [{"term": "(none)", "growth": None}])
        rq30 = d30.get("risingQueries", [{"term": "(none)", "growth": None}])
        for i in range(max(len(rq90), len(rq30), 1)):
            r90 = rq90[i] if i < len(rq90) else {}
            r30 = rq30[i] if i < len(rq30) else {}
            ws.append([r90.get("term", ""), r90.get("growth"), None, None, r30.get("term", ""), r30.get("growth")])

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    wb.save(output_path)
    log(f"Excel saved → {output_path}")


# ── JS writer ─────────────────────────────────────────────────────────────────
def write_js(all_data, pulled_at, js_path):
    now_iso = pulled_at.strftime("%Y-%m-%dT%H:%M:%SZ")
    date_str = pulled_at.strftime("%Y-%m-%d")

    lines = [
        "// Google Trends Data — Crocs",
        f"// Auto-generated by fetch_trends.py on {date_str}",
        f"// DO NOT EDIT MANUALLY — re-run fetch_trends.py to update",
        "",
        "const GTRENDS_META = {",
        f'  keyword: "{KEYWORD}",',
        f'  pulledAt: "{now_iso}",',
        '  note: "Interest is normalized 0–100 within each market+timeframe — NOT comparable across markets in absolute terms.",',
        "};",
        "",
        "const GTRENDS_DATA = {",
    ]

    for market_name, geo in MARKETS:
        mdata = all_data.get(market_name, {})
        d90 = mdata.get("90d", {})
        d30 = mdata.get("30d", {})

        def tf_block(d, tf):
            if not d or d.get("blocked"):
                return (
                    f"    '{tf}': {{\n"
                    f"      avgInterest: null, peakInterest: null,\n"
                    f"      interestOverTime: [], topQueries: [], risingQueries: [],\n"
                    f"      blocked: true,\n"
                    f"      blockedNote: 'Google Trends data is unavailable for this market.',\n"
                    f"    }},"
                )
            iot = json.dumps(d.get("interestOverTime", []))
            tq = json.dumps(d.get("topQueries", []))
            rq = json.dumps(d.get("risingQueries", []))
            avg = d.get("avgInterest")
            peak = d.get("peakInterest")
            avg_str = str(avg) if avg is not None else "null"
            peak_str = str(peak) if peak is not None else "null"
            return (
                f"    '{tf}': {{\n"
                f"      avgInterest: {avg_str}, peakInterest: {peak_str},\n"
                f"      interestOverTime: {iot},\n"
                f"      topQueries: {tq},\n"
                f"      risingQueries: {rq},\n"
                f"    }},"
            )

        lines.append(f"  '{market_name}': {{")
        lines.append(f"    geo: '{geo}',")
        lines.append(tf_block(d90, "90d"))
        lines.append(tf_block(d30, "30d"))
        lines.append("  },")

    lines.append("};")
    lines.append("")
    lines.append("const GTRENDS_MARKETS = [")
    lines.append("  " + ", ".join(f"'{m}'" for m, _ in MARKETS))
    lines.append("];")
    lines.append("")
    lines.append("function getGtrendsSummary() {")
    lines.append("  return GTRENDS_MARKETS.map(market => {")
    lines.append("    const mdata = GTRENDS_DATA[market];")
    lines.append("    return {")
    lines.append("      market, geo: mdata.geo, blocked: !!(mdata['90d'] && mdata['90d'].blocked),")
    lines.append("      avg90d: mdata['90d']?.avgInterest, peak90d: mdata['90d']?.peakInterest,")
    lines.append("      rising90d: (mdata['90d']?.risingQueries||[]).length,")
    lines.append("      breakout90d: (mdata['90d']?.risingQueries||[]).filter(q=>q.type==='breakout').length,")
    lines.append("      avg30d: mdata['30d']?.avgInterest, peak30d: mdata['30d']?.peakInterest,")
    lines.append("      rising30d: (mdata['30d']?.risingQueries||[]).length,")
    lines.append("      breakout30d: (mdata['30d']?.risingQueries||[]).filter(q=>q.type==='breakout').length,")
    lines.append("    };")
    lines.append("  });")
    lines.append("}")

    with open(js_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    log(f"JS data file saved → {js_path}")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Fetch Google Trends data for Crocs")
    parser.add_argument("--api", choices=["pytrends", "serpapi"], default="pytrends",
                        help="Which API to use (default: pytrends)")
    parser.add_argument("--key", default=None,
                        help="SerpAPI key (required if --api serpapi)")
    parser.add_argument("--cookies", action="store_true",
                        help="Use Chrome cookies to help bypass pytrends blocking")
    parser.add_argument("--out-dir", default=OUTPUT_DIR,
                        help="Output directory (default: same as this script)")
    args = parser.parse_args()

    if args.api == "serpapi" and not args.key:
        print("ERROR: --key is required when using --api serpapi")
        print("  Sign up free at serpapi.com (100 free searches/month)")
        sys.exit(1)

    pulled_at = datetime.datetime.utcnow()
    all_data = {}

    log(f"Starting Crocs Google Trends fetch via {args.api.upper()}")
    log(f"Keyword: {KEYWORD} | Markets: {len(MARKETS)} | Timeframes: {len(TIMEFRAMES)}")
    log("")

    for market_name, geo in MARKETS:
        # China special case — Google is blocked there
        if geo == "CN":
            log(f"Skipping {market_name} (CN) — Google Trends not available in mainland China")
            all_data[market_name] = {
                "90d": {"blocked": True, "blockedNote": "Google services are blocked in mainland China.",
                        "avgInterest": None, "peakInterest": None,
                        "interestOverTime": [], "topQueries": [], "risingQueries": []},
                "30d": {"blocked": True, "blockedNote": "Google services are blocked in mainland China.",
                        "avgInterest": None, "peakInterest": None,
                        "interestOverTime": [], "topQueries": [], "risingQueries": []},
            }
            continue

        all_data[market_name] = {}
        log(f"Market: {market_name} ({geo})")
        for tf_label, tf_google in TIMEFRAMES:
            if args.api == "serpapi":
                result = fetch_with_serpapi(market_name, geo, tf_label, tf_google, args.key)
            else:
                result = fetch_with_pytrends(market_name, geo, tf_label, tf_google, use_cookies=args.cookies)
            all_data[market_name][tf_label] = result
            log(f"  → {tf_label}: avg={result.get('avgInterest')}, peak={result.get('peakInterest')}, "
                f"rising={len(result.get('risingQueries', []))}")

        sleep_politely(3, 6)  # extra pause between markets

    date_str = pulled_at.strftime("%Y-%m-%d")
    xlsx_path = os.path.join(args.out_dir, f"crocs_trends_{date_str}.xlsx")
    js_path = os.path.join(args.out_dir, "google-trends-data.js")

    log("")
    log("Writing outputs …")
    write_excel(all_data, pulled_at, xlsx_path)
    write_js(all_data, pulled_at, js_path)

    log("")
    log("Done! Next steps:")
    log("  1. Upload the Excel to the dashboard (drag & drop) — or")
    log("  2. git add google-trends-data.js && git commit -m 'Update Google Trends data' && git push")
    log("  3. The dashboard at GitHub Pages will reflect updated data within ~60 seconds")


if __name__ == "__main__":
    main()
